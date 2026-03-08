#!/usr/bin/env python3
"""
sync_lockring_from_excel.py

Usage:
  python sync_lockring_from_excel.py --excel path/to/lockring.xlsx [--html path/to/lockring-search.html]

This script reads an Excel file and converts rows into a JS object assigned to `const excelData`
inside the target HTML. It is flexible about header names and groups rows by a "group" column
(which it will try to detect). Fields it looks for (case-insensitive):
  - group (or no/index/category) -> used as the key in the outer object
  - whereToUse (where, usage)
  - partNumber (part, partnumber)
  - connectorName (connector, connectorname)
  - usageArea (usageArea, usage)
  - pipeSize (pipe, pipesize, size)

If headers are missing, it will attempt reasonable fallbacks and will warn.
"""

import argparse
import json
import os
import re
import sys

try:
    import openpyxl
except Exception as e:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    raise


def normalize(s):
    return re.sub(r"[^a-z0-9]", "", s.lower()) if s else ""


def detect_column(headers, candidates):
    # headers: list of original header strings
    # candidates: list of candidate tokens to match
    for i, h in enumerate(headers):
        nh = normalize(str(h))
        for c in candidates:
            if c in nh:
                return i
    return None


def read_sheet(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    # choose first sheet
    sheet = wb[wb.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(c) if c is not None else "" for c in rows[0]]
    data_rows = rows[1:]
    return headers, data_rows


def build_excel_data(headers, rows):
    # find group column
    group_idx = detect_column(headers, ["group", "no", "index", "category", "key"])
    where_idx = detect_column(headers, ["wheretouse", "where", "usage", "usagetype"])
    part_idx = detect_column(headers, ["partnumber", "partnumber", "part", "part_num", "partnum"])
    connector_idx = detect_column(headers, ["connectorname", "connector", "connect"])
    usagearea_idx = detect_column(headers, ["usagearea", "usagearea", "usearea", "usage"])
    pipesize_idx = detect_column(headers, ["pipesize", "pipe", "size", "pipe_size"]) 

    # If no group column, use first column
    if group_idx is None:
        group_idx = 0

    result = {}
    warnings = []

    for r in rows:
        key_raw = r[group_idx] if group_idx < len(r) else None
        if key_raw is None:
            continue
        key = str(key_raw).strip()
        if key == "":
            continue
        item = {}
        def get(i):
            try:
                return r[i] if i is not None and i < len(r) else None
            except Exception:
                return None

        item['whereToUse'] = str(get(where_idx)).strip() if where_idx is not None and get(where_idx) is not None else ""
        item['partNumber'] = str(get(part_idx)).strip() if part_idx is not None and get(part_idx) is not None else ""
        item['connectorName'] = str(get(connector_idx)).strip() if connector_idx is not None and get(connector_idx) is not None else ""
        item['usageArea'] = str(get(usagearea_idx)).strip() if usagearea_idx is not None and get(usagearea_idx) is not None else ""
        item['pipeSize'] = str(get(pipesize_idx)).strip() if pipesize_idx is not None and get(pipesize_idx) is not None else ""

        result.setdefault(str(key), []).append(item)

    # Basic warnings
    if part_idx is None:
        warnings.append('partNumber column not detected; output will have empty partNumber fields.')
    if where_idx is None:
        warnings.append('whereToUse column not detected; output will have empty whereToUse fields.')

    return result, warnings


def replace_in_html(html_path, new_js_text, dry_run=False):
    with open(html_path, 'r', encoding='utf-8') as f:
        content = f.read()

    pattern = re.compile(r"const\s+excelData\s*=\s*\{.*?\n\s*\};", re.DOTALL)
    if not pattern.search(content):
        print("Warning: existing `const excelData = {...};` block not found in HTML. Appending new block before </script> if possible.")
        # try to insert before closing </script> of the first script that contains excelData or before </body>
        insert_point = content.rfind('</script>')
        if insert_point == -1:
            insert_point = content.rfind('</body>')
        if insert_point == -1:
            insert_point = len(content)
        new_content = content[:insert_point] + "\n" + new_js_text + "\n" + content[insert_point:]
    else:
        new_content = pattern.sub(new_js_text, content)

    if dry_run:
        print('Dry run: not writing changes to', html_path)
        return True

    # write with utf-8-sig to be safe for BOM if needed
    with open(html_path, 'w', encoding='utf-8-sig') as f:
        f.write(new_content)
    return True


def main():
    p = argparse.ArgumentParser()
    p.add_argument('--excel', '-e', required=True, help='Path to updated Excel (.xlsx) file')
    p.add_argument('--html', '-h', default='lockring-search.html', help='Path to lockring-search.html')
    p.add_argument('--dry-run', action='store_true')
    args = p.parse_args()

    excel_path = os.path.abspath(args.excel)
    html_path = os.path.abspath(args.html)

    if not os.path.exists(excel_path):
        print(f'Error: excel file not found: {excel_path}')
        sys.exit(1)
    if not os.path.exists(html_path):
        print(f'Error: html file not found: {html_path}')
        sys.exit(1)

    headers, rows = read_sheet(excel_path)
    data, warnings = build_excel_data(headers, rows)

    if warnings:
        print('Warnings:')
        for w in warnings:
            print(' -', w)

    # Build JS text
    json_text = json.dumps(data, ensure_ascii=False, indent=2)
    # ensure double quotes consistent with examples
    new_js = 'const excelData = ' + json_text + ';'

    print('Generated excelData for', len(data), 'groups.')

    ok = replace_in_html(html_path, new_js, dry_run=args.dry_run)
    if ok:
        if args.dry_run:
            print('Dry run succeeded. No file written.')
        else:
            print('HTML updated successfully:', html_path)


if __name__ == '__main__':
    main()
